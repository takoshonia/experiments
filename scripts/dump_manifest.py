#!/usr/bin/env python3
"""Extract a sample from the STT audio manifest xlsx into a JSONL file.

Usage (Windows):
  python scripts/dump_manifest.py ^
    --manifest "C:\\Users\\tamar\\Downloads\\TRAINING_PROCESSED_STT_audio_manifest\\TRAINING_PROCESSED_STT_audio_manifest.xlsx" ^
    --sample-size 100 ^
    --seed 42 ^
    --output manifest_sample.jsonl

Output JSONL row:
  {"stem": "...", "reference": "...", "prediction": "...", "wer_before": 0.42, ...}
"""

from __future__ import annotations

import argparse
import json
import random
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: pip install openpyxl"
    ) from exc


REQUIRED_COLUMNS = ("folder", "filename", "extension", "Text", "STT1_Text")
OPTIONAL_COLUMNS = ("Duration", "WER", "CER", "BLEU")


def _read_rows(manifest_path: Path) -> tuple[list[str], list[tuple]]:
    wb = load_workbook(manifest_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    rows = [r for r in rows_iter if r and any(c is not None for c in r)]
    wb.close()
    return header, rows


def _index_columns(header: list[str]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for col in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
        if col not in header:
            if col in REQUIRED_COLUMNS:
                raise SystemExit(f"Required column missing in manifest: {col}")
            continue
        cols[col] = header.index(col)
    return cols


def _row_to_dict(row: tuple, cols: dict[str, int]) -> dict | None:
    folder = row[cols["folder"]] or ""
    filename = row[cols["filename"]] or ""
    ext = row[cols["extension"]] or ""
    reference = (row[cols["Text"]] or "").strip()
    prediction = (row[cols["STT1_Text"]] or "").strip()
    if not reference or not prediction:
        return None
    stem = f"{folder}/{filename}{ext}" if folder else f"{filename}{ext}"
    item = {
        "stem": stem,
        "reference": reference,
        "prediction": prediction,
    }
    if "Duration" in cols and row[cols["Duration"]] is not None:
        item["duration_s"] = float(row[cols["Duration"]])
    if "WER" in cols and row[cols["WER"]] is not None:
        item["wer_before"] = float(row[cols["WER"]])
    if "CER" in cols and row[cols["CER"]] is not None:
        item["cer_before"] = float(row[cols["CER"]])
    if "BLEU" in cols and row[cols["BLEU"]] is not None:
        item["bleu_before"] = float(row[cols["BLEU"]])
    return item


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", type=Path, required=True, help="Path to the xlsx manifest.")
    parser.add_argument("--sample-size", type=int, default=100, help="Number of rows to sample (default: 100).")
    parser.add_argument("--seed", type=int, default=42, help="RNG seed for reproducible sampling.")
    parser.add_argument("--output", type=Path, default=Path("manifest_sample.jsonl"), help="Output JSONL path.")
    parser.add_argument(
        "--min-words",
        type=int,
        default=3,
        help="Skip rows whose reference has fewer than N words (default: 3).",
    )
    args = parser.parse_args()

    if not args.manifest.is_file():
        raise SystemExit(f"Manifest not found: {args.manifest}")

    print(f"Reading {args.manifest} ...", flush=True)
    header, rows = _read_rows(args.manifest)
    print(f"Loaded header: {header}", flush=True)
    print(f"Data rows: {len(rows):,}", flush=True)

    cols = _index_columns(header)
    parsed: list[dict] = []
    for raw in rows:
        item = _row_to_dict(raw, cols)
        if not item:
            continue
        if len(item["reference"].split()) < args.min_words:
            continue
        parsed.append(item)
    print(f"Usable rows (with text + prediction, >= {args.min_words} words): {len(parsed):,}", flush=True)

    if args.sample_size < len(parsed):
        rng = random.Random(args.seed)
        sample = rng.sample(parsed, args.sample_size)
    else:
        sample = parsed

    with args.output.open("w", encoding="utf-8") as f:
        for item in sample:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    print(f"Wrote {len(sample)} rows -> {args.output}", flush=True)


if __name__ == "__main__":
    main()
